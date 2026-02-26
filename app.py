import streamlit as st
import openpyxl
import tempfile
import shutil
import os

ARCHIVO = "simulador.xlsx"
NOMBRE_HOJA = "versf"

st.set_page_config(page_title="Simulador CFG", layout="centered")

st.title("📊 Simulador Comercial")

# ================================
# VERIFICAR QUE EL ARCHIVO EXISTE
# ================================

if not os.path.exists(ARCHIVO):
    st.error(f"No se encuentra el archivo {ARCHIVO}")
    st.write("Archivos disponibles en la carpeta:")
    st.write(os.listdir())
    st.stop()

# ================================
# CAMPOS DE ENTRADA
# ================================

col1, col2 = st.columns(2)

with col1:
    ciudad = st.text_input("Ciudad")
    cuota_oro = st.number_input("Cuota ventas oro", min_value=0.0)
    cuota_semestre = st.number_input("Cuota Pago anticipado Semestre", min_value=0.0)
    cuota_anio = st.number_input("Cuota pago anticipado año", min_value=0.0)
    cuota_trimestre = st.number_input("Cuota pago anticipado trimestre", min_value=0.0)
    cuota_plan = st.number_input("Cuota plan asociado", min_value=0.0)

with col2:
    tipo_colectivo = st.text_input("Tipo colectivo")
    cuota_colectivo = st.number_input("Cuota colectivo", min_value=0.0)
    tiempo = st.text_input("Tiempo")
    numero_ventas = st.number_input("# Número de ventas", min_value=0)
    acelerador = st.number_input("Acelerador plan asociado", min_value=0.0)

# ================================
# BOTÓN CALCULAR
# ================================

if st.button("🚀 Calcular"):

    try:
        # Crear copia temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            shutil.copy(ARCHIVO, tmp.name)

            wb = openpyxl.load_workbook(tmp.name)
            ws = wb[NOMBRE_HOJA]

            # Escribir entradas
            ws["C3"] = ciudad
            ws["F3"] = cuota_oro
            ws["G3"] = cuota_semestre
            ws["H3"] = cuota_anio
            ws["I3"] = cuota_trimestre
            ws["J3"] = cuota_plan
            ws["K3"] = tipo_colectivo
            ws["L3"] = cuota_colectivo
            ws["O3"] = tiempo
            ws["U3"] = numero_ventas
            ws["B7"] = acelerador

            wb.save(tmp.name)
            wb.close()

            # Reabrir para leer cálculos
            wb_calc = openpyxl.load_workbook(tmp.name, data_only=True)
            ws_calc = wb_calc[NOMBRE_HOJA]

            st.subheader("📈 Resultados")

            def formato_numero(valor):
                if isinstance(valor, (int, float)):
                    return f"{valor:,.0f}".replace(",", ".")
                return valor

            st.success(f"Cuota cumplimiento: {formato_numero(ws_calc['E3'].value)}")
            st.success(f"Cuanto se vendió: {formato_numero(ws_calc['M3'].value)}")
            st.success(f"Cuanto tengo: {formato_numero(ws_calc['N3'].value)}")

            valor_q3 = ws_calc["Q3"].value
            if isinstance(valor_q3, (int, float)):
                if valor_q3 <= 1:
                    valor_q3 = f"{valor_q3:.1%}"
                else:
                    valor_q3 = f"{valor_q3:.1f}%"

            st.success(f"Cumplimiento comisión: {valor_q3}")
            st.success(f"Te falta: {formato_numero(ws_calc['R3'].value)}")
            st.success(f"# te falta: {ws_calc['T3'].value}")

            wb_calc.close()

    except Exception as e:
        st.error(f"Error: {e}")